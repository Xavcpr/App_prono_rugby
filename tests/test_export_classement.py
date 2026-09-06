import os
import pytest

from django.contrib.auth.models import User
from core.models import (
    Competition, Season, Team, Player,
    CompetitionTeamPrediction, CompetitionBonusPrediction,
)
from django.core.management import call_command


@pytest.fixture
def export_data(competition, season):
    teams = [Team.objects.create(name=f"Équipe {i}") for i in range(14)]
    players = []
    for name in ["Alice", "Bob", "Charlie"]:
        u = User.objects.create_user(username=f"u-{name.lower()}", password="x")
        p = Player.objects.create(user=u, name=name)
        players.append(p)
        for pos, team in enumerate(teams, 1):
            CompetitionTeamPrediction.objects.create(
                player=p, season=season, competition=competition,
                team=team, position=pos,
            )
        CompetitionBonusPrediction.objects.create(
            player=p, season=season, competition=competition,
            winner=teams[0],
            best_try_scorer="Dupont",
            best_point_scorer="Ntamack",
        )
    return players, teams


@pytest.mark.django_db
def test_export_classement_creates_valid_xlsx(export_data, competition, season, tmp_path):
    out = str(tmp_path / "test_export.xlsx")
    call_command(
        "export_classement",
        competition=str(competition.id),
        season=str(season.id),
        output=out,
    )
    assert os.path.exists(out)

    import openpyxl
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    players, teams = export_data

    # 3 players → cols B, D, F
    # Labels in A
    assert ws.cell(1, 1).value == "Joueur"
    assert ws.cell(2, 1).value == "1er"
    assert ws.cell(15, 1).value == "14ème"
    assert ws.cell(16, 1).value == "Vainqueur"
    assert ws.cell(17, 1).value == "Meilleur marqueur"
    assert ws.cell(18, 1).value == "Meilleur réalisateur"

    # Player names in row 1
    assert ws.cell(1, 2).value == "Alice"   # B1
    assert ws.cell(1, 4).value == "Bob"     # D1
    assert ws.cell(1, 6).value == "Charlie" # F1

    # Positions
    assert ws.cell(2, 2).value == "Équipe 0"  # Alice, pos 1
    assert ws.cell(15, 2).value == "Équipe 13" # Alice, pos 14

    # Bonus
    assert ws.cell(16, 2).value == "Équipe 0"  # winner
    assert ws.cell(17, 2).value == "Dupont"    # try scorer
    assert ws.cell(18, 2).value == "Ntamack"   # point scorer

    # Column C is empty (separator)
    assert ws.cell(1, 3).value is None
