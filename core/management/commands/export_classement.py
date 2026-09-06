import sys

from django.core.management.base import BaseCommand, CommandError
from django.db.models import Min

from core.models import (
    Competition, Season, Player,
    CompetitionTeamPrediction, CompetitionBonusPrediction,
)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    raise CommandError("openpyxl is required: pip install openpyxl")


THIN = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
BOLD   = Font(bold=True, size=11)
HDR_FONT = Font(bold=True, size=11)


class Command(BaseCommand):
    help = "Exporte le classement pronos (positions + bonus) vers Excel"

    def add_arguments(self, parser):
        parser.add_argument("-c", "--competition", required=True,
                            help="ID de la compétition")
        parser.add_argument("-s", "--season", required=True,
                            help="ID de la saison")
        parser.add_argument("-o", "--output", default="classement_pronos.xlsx",
                            help="Nom du fichier de sortie")

    def handle(self, *args, **options):
        try:
            competition = Competition.objects.get(id=int(options["competition"]))
            season = Season.objects.get(id=int(options["season"]))
        except (Competition.DoesNotExist, Season.DoesNotExist) as exc:
            raise CommandError(str(exc))

        # ── Joueurs (ordre alpha, même logique que recap_pronos) ──────
        player_ids = (
            CompetitionTeamPrediction.objects
            .filter(competition=competition, season=season)
            .values_list("player_id", flat=True)
            .distinct()
        )
        players = list(
            Player.objects.filter(id__in=player_ids).order_by("name")
        )
        if not players:
            self.stdout.write(self.style.WARNING("Aucun joueur avec pronos."))
            return

        # ── Prédictions d'équipe : {player_id: {position: team_name}} ─
        team_preds = {}
        for p in (
            CompetitionTeamPrediction.objects
            .filter(competition=competition, season=season)
            .select_related("team")
        ):
            team_preds.setdefault(p.player_id, {})[p.position] = p.team.name

        # ── Bonus predictions : {player_id: {winner, try, point}} ────
        bonus_preds = {}
        for bp in (
            CompetitionBonusPrediction.objects
            .filter(competition=competition, season=season)
            .select_related("winner")
        ):
            bonus_preds[bp.player_id] = {
                "winner":  bp.winner.name if bp.winner else "",
                "try":     bp.best_try_scorer or "",
                "point":   bp.best_point_scorer or "",
            }

        # ── Nombre max de positions ───────────────────────────────────
        max_pos = max(
            (pos for positions in team_preds.values() for pos in positions),
            default=14,
        )

        # ── Labels colonne A ──────────────────────────────────────────
        labels = ["Joueur"]
        for i in range(1, max_pos + 1):
            suffix = {1: "er", 2: "ème"}.get(i, "ème") if i <= 2 else "ème"
            labels.append(f"{i}{suffix}")
        labels += ["Vainqueur", "Meilleur marqueur", "Meilleur réalisateur"]
        n_labels = len(labels)  # 1 + max_pos + 3

        # ── Création du workbook ──────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Classement pronos"

        # ── Écriture des données ──────────────────────────────────────
        for pi, player in enumerate(players):
            col = 2 + pi * 2           # B=2, D=4, F=6, … (impair = données)
            preds = team_preds.get(player.id, {})
            bonus = bonus_preds.get(player.id, {})

            for row_idx in range(1, n_labels + 1):
                cell = ws.cell(row=row_idx, column=col)
                if row_idx == 1:
                    cell.value = player.name
                elif row_idx <= max_pos + 1:
                    cell.value = preds.get(row_idx - 1, "")
                elif row_idx == max_pos + 2:
                    cell.value = bonus.get("winner", "")
                elif row_idx == max_pos + 3:
                    cell.value = bonus.get("try", "")
                elif row_idx == max_pos + 4:
                    cell.value = bonus.get("point", "")

                cell.border = THIN
                cell.alignment = CENTER if row_idx != 1 else LEFT

        # ── Colonne A (labels) ────────────────────────────────────────
        for row_idx, label in enumerate(labels, 1):
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.font = BOLD
            cell.border = THIN
            cell.alignment = LEFT

        # ── Mise en forme ─────────────────────────────────────────────
        ws.column_dimensions["A"].width = 20
        for pi in range(len(players)):
            col_letter = openpyxl.utils.get_column_letter(2 + pi * 2)
            ws.column_dimensions[col_letter].width = 22

        ws.freeze_panes = "B2"

        out = options["output"]
        wb.save(out)
        self.stdout.write(self.style.SUCCESS(
            f"{out} — {len(players)} joueurs, {max_pos} positions"
        ))
