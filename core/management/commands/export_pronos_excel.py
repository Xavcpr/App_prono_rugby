import logging
from collections import defaultdict

from django.core.management.base import BaseCommand, CommandError
from django.db.models import Sum

from core.models import (
    Competition, Season, Round, Match, Player,
    Prediction, DailyScore
)
from core.services.scoring import _get_scoring_config, PHASE_MULTIPLIERS, BONUS_SCALES

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise CommandError("openpyxl is required: pip install openpyxl")

logger = logging.getLogger(__name__)

# ---- Styles ----
HDR_FILL = PatternFill("solid", fgColor="4472C4")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
GRN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
BLU_FILL = PatternFill("solid", fgColor="BDD7EE")
YEL_FILL = PatternFill("solid", fgColor="FFFFCC")
TOT_FILL = PatternFill("solid", fgColor="FF6B6B")
GRN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006")
BLU_FONT = Font(color="1F4E79", bold=True, size=12)
BOLD = Font(bold=True, size=11)
THIN = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

def _c(ws, r, c, v=None, font=None, fill=None, align=None, border=THIN):
    cell = ws.cell(row=r, column=c, value=v)
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    cell.border = border
    return cell

def _merge(ws, r1, r2, c1, c2, v=None, font=None, fill=None, align=None):
    if r1 != r2 or c1 != c2:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = _c(ws, r1, c1, v, font=font, fill=fill, align=align)
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            c_obj = ws.cell(row=rr, column=cc)
            if fill: c_obj.fill = fill
            c_obj.border = THIN
    return cell


class Command(BaseCommand):
    help = "Exporte pronostics + scores vers Excel structuré"

    def add_arguments(self, parser):
        parser.add_argument("-o", "--output", default="export_pronos.xlsx")
        parser.add_argument("-c", "--competition", help="Nom compétition (optionnel)")
        parser.add_argument("-s", "--season", help="Année saison (ex: 2026/2027)")

    def handle(self, *args, **options):
        out = options["output"]
        qs = Season.objects.filter(year__gte=2025).order_by("-year")
        if options["competition"]:
            qs = qs.filter(competition__name__icontains=options["competition"])
        if options["season"]:
            qs = qs.filter(year=options["season"])

        seasons = list(qs)
        if not seasons:
            self.stdout.write(self.style.WARNING("Aucune saison."))
            return

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        total_rounds = 0
        for season in seasons:
            rounds = Round.objects.filter(season=season).order_by("number")
            for rnd in rounds:
                self._build_sheet(wb, f"J{rnd.number}", rnd, season)
                total_rounds += 1
        wb.save(out)
        self.stdout.write(self.style.SUCCESS(f"{out} — {total_rounds} journées"))

    # ------------------------------------------------------------------
    def _build_sheet(self, wb, name, rnd, season):
        ws = wb.create_sheet(title=name[:31])
        comp = season.competition
        phase = rnd.phase or "POOL"
        cfg = _get_scoring_config(season)["SCORING_CONFIG"]
        mult = PHASE_MULTIPLIERS.get(phase, 1.0)
        scale = BONUS_SCALES.get(comp.name, {})

        matches = list(Match.objects.filter(round=rnd).select_related("home_team","away_team").order_by("kickoff_at"))
        n_matches = len(matches)

        players = list(Player.objects.filter(seasons=season).order_by("name"))
        n_players = len(players)

        all_preds = Prediction.objects.filter(match__round=rnd).select_related("player","match")
        pred_idx = {}
        for pr in all_preds:
            pred_idx[(pr.player_id, pr.match_id)] = pr

        daily_pts = {ds.user_id: ds.points for ds in DailyScore.objects.filter(round=rnd)}

        # -- points déjà stockés via process_round_scores --
        pts_by_player = defaultdict(int)
        correct_by_player = defaultdict(int)
        bonus_off = defaultdict(int)
        bonus_def = defaultdict(int)
        bonus_diff = defaultdict(int)
        bonus_sum = defaultdict(int)
        bonus_away = defaultdict(int)
        half_full = defaultdict(int)
        match_pts_detail = defaultdict(dict)

        for pr in all_preds:
            pid = pr.player_id
            pts_by_player[pid] += pr.points or 0
            mid = pr.match_id
            m = next((x for x in matches if x.id == mid), None)
            if not m or m.home_score is None:
                continue
            real_side = (m.home_score > m.away_score and "HOME") or (m.away_score > m.home_score and "AWAY") or "DRAW"
            pred_side = (pr.home_score_pred > pr.away_score_pred and "HOME") or (pr.away_score_pred > pr.home_score_pred and "AWAY") or "DRAW"
            if real_side == pred_side:
                correct_by_player[pid] += 1

        # Calcul des bonus individuels via la logique « à la main » pour l'affichage
        # (On lit depuis les DailyScore déjà écrits pour le score final)
        from core.services.scoring import calculate_match_points
        match_val_map = {}
        for m in matches:
            if m.home_score is not None:
                winners_cnt = sum(1 for pr in all_preds if pr.match_id == m.id and
                                  pr.home_score_pred is not None and
                                  ((m.home_score > m.away_score and pr.home_score_pred > pr.away_score_pred) or
                                   (m.away_score > m.home_score and pr.away_score_pred > pr.home_score_pred) or
                                   (m.home_score == m.away_score and pr.home_score_pred == pr.away_score_pred)))
                base = cfg["MATCH_POOL_BASE"]
                perfect = cfg["PERFECT_SCORE_BONUS"]
                match_val_map[m.id] = int((base * mult + perfect) / max(winners_cnt, 1))
            else:
                match_val_map[m.id] = 0

        # --- TABLEAU BAS d'abord (matchs + pronos) ---
        B = 20
        fix_headers = ["Domicile","vs","Extérieur","Nb D","BO D réel","Score D","-","Score E","BO E réel","","Valeur",""]
        for ci, h in enumerate(fix_headers, 1):
            _c(ws, B, ci, h, font=HDR_FONT, fill=HDR_FILL, align=CENTER)

        # En-têtes joueurs
        p_cols = []
        col = 13
        for p in players:
            abbr = p.name.split()[0][:8]
            p_cols.append((col, col + 5))
            _c(ws, B, col, abbr, font=HDR_FONT, fill=HDR_FILL, align=CENTER)
            for off in range(1, 6):
                ws.cell(row=B, column=col + off).fill = HDR_FILL
                ws.cell(row=B, column=col + off).border = THIN
            col += 6

        # Matchs
        for mi, m in enumerate(matches):
            row = B + 1 + mi
            if row > B + 12:
                break
            _c(ws, row, 1, m.home_team.name if m.home_team else "", align=RIGHT)
            _c(ws, row, 2, "-", align=CENTER)
            _c(ws, row, 3, m.away_team.name if m.away_team else "", align=LEFT)
            n_pred_dom = sum(1 for pr in all_preds if pr.match_id == m.id and pr.home_score_pred is not None and pr.home_score_pred > pr.away_score_pred)
            _c(ws, row, 4, n_pred_dom if n_pred_dom else "", align=CENTER)
            if m.home_score is not None:
                _c(ws, row, 5, "X" if m.bonus_offense_home else "", align=CENTER)
                _c(ws, row, 6, m.home_score, align=CENTER)
                _c(ws, row, 7, "-", align=CENTER)
                _c(ws, row, 8, m.away_score, align=CENTER)
                _c(ws, row, 9, "X" if m.bonus_offense_away else "", align=CENTER)
            _c(ws, row, 11, match_val_map.get(m.id, 0) if m.home_score is not None else "", align=CENTER)

            for pi, pl in enumerate(players):
                sc_start = p_cols[pi][0]
                pr = pred_idx.get((pl.id, m.id))
                if pr and pr.home_score_pred is not None and m.home_score is not None:
                    boh, boa = pr.bonus_home_pred, pr.bonus_away_pred
                    boh_ok = boh and m.bonus_offense_home
                    boa_ok = boa and m.bonus_offense_away
                    _c(ws, row, sc_start, "X" if boh else "",
                       font=GRN_FONT if boh_ok else (RED_FONT if boh else None),
                       fill=GRN_FILL if boh_ok else (RED_FILL if boh and not boh_ok else None), align=CENTER)
                    _c(ws, row, sc_start + 1, pr.home_score_pred,
                       fill=GRN_FILL if pr.home_score_pred == m.home_score else None,
                       font=GRN_FONT if pr.home_score_pred == m.home_score else None, align=CENTER)
                    _c(ws, row, sc_start + 2, "-", align=CENTER)
                    _c(ws, row, sc_start + 3, pr.away_score_pred,
                       fill=GRN_FILL if pr.away_score_pred == m.away_score else None,
                       font=GRN_FONT if pr.away_score_pred == m.away_score else None, align=CENTER)
                    _c(ws, row, sc_start + 4, "X" if boa else "",
                       font=GRN_FONT if boa_ok else (RED_FONT if boa else None),
                       fill=GRN_FILL if boa_ok else (RED_FILL if boa and not boa_ok else None), align=CENTER)
                    _c(ws, row, sc_start + 5, pr.points or 0, font=BLU_FONT, fill=BLU_FILL, align=CENTER)
                else:
                    for off in range(6):
                        _c(ws, row, sc_start + off, "", align=CENTER)

        # Totaux
        tr = B + 1 + n_matches + 1
        _c(ws, tr, 1, "TOTAUX", font=BOLD, fill=YEL_FILL, align=RIGHT)
        for pi, pl in enumerate(players):
            ec = p_cols[pi][1]
            first_r = B + 1
            last_r = B + n_matches
            formula = f"=SUM({get_column_letter(ec)}{first_r}:{get_column_letter(ec)}{last_r})"
            _c(ws, tr, ec, formula, font=Font(bold=True, color="9C0006", size=11), fill=TOT_FILL, align=CENTER)

        for r in range(B + 1 + n_matches, B + 13):
            ws.row_dimensions[r].hidden = True

        # --- TABLEAU HAUT (classement journée) — écrit APRÈS le tableau bas ---
        hdr_top = [
            (24, "Rang"), (25, 29, "Joueur"), (30, 32, "Score J"),
            (33, 34, "Bons"), (35, 36, "Bonus J"),
            (37, 38, "Pts bons"), (39, 40, "Pts ½/Pile"),
            (41, 42, "Pts Off"), (43, 44, "Pts Def"),
            (45, 46, "Pts Diff"), (47, 48, "Pts Somme"),
            (49, 50, "Pts Away"),
        ]
        for h in hdr_top:
            if len(h) == 2:
                _c(ws, 1, h[0], h[1], font=HDR_FONT, fill=HDR_FILL, align=CENTER)
            else:
                _merge(ws, 1, 1, h[0], h[1], h[2], font=HDR_FONT, fill=HDR_FILL, align=CENTER)

        sorted_players = sorted(players, key=lambda p: -(daily_pts.get(p.user_id, 0)))
        for i, p in enumerate(sorted_players):
            row = i + 2
            total_pts = daily_pts.get(p.user_id, 0)
            n_correct = correct_by_player.get(p.id, 0)
            day_bonus = 0
            for t in sorted(scale.keys(), reverse=True):
                if n_correct >= t:
                    day_bonus = scale[t]
                    break
            fill = ALT_FILL if i % 2 == 0 else None
            _c(ws, row, 24, i + 1, align=CENTER)
            _merge(ws, row, row, 25, 29, p.name, fill=fill, align=LEFT)
            _merge(ws, row, row, 30, 32, total_pts, font=BLU_FONT, fill=fill, align=CENTER)
            _merge(ws, row, row, 33, 34, f"{n_correct}/{n_matches}", fill=fill, align=CENTER)
            _merge(ws, row, row, 35, 36, day_bonus if day_bonus else "", fill=fill, align=CENTER)
            for c1, c2 in [(37, 38), (39, 40), (41, 42), (43, 44), (45, 46), (47, 48), (49, 50)]:
                for cc in range(c1, c2 + 1):
                    _c(ws, row, cc, "", fill=fill)

        # Largeurs & gel
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 4
        ws.column_dimensions["C"].width = 16
        for cl in "DEFGHIK":
            ws.column_dimensions[cl].width = 8
        ws.column_dimensions["J"].width = 3
        ws.column_dimensions["L"].width = 3
        ws.freeze_panes = "B2"
